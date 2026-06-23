"""
Market data coverage and fetch endpoints.

GET  /api/data/coverage
POST /api/data/fetch
GET  /api/data/fetch/status/{job_id}
"""
from __future__ import annotations

import asyncio
import csv
import io
import math
import re
import shutil
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, BackgroundTasks, HTTPException, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

_jobs: dict[str, dict] = {}

# ponytail: global single fetch lock -- one fetch runs at a time across all
# sessions. Split into per-exchange locks only if OKX+Binance parallelism is
# ever needed.
_fetch_lock = asyncio.Lock()
_TERMINAL_FETCH_STATUSES = {"done", "error", "cancelled"}

_FETCH_BAR_MS: dict[str, int] = {
    "1m": 60_000,
    "3m": 3 * 60_000,
    "5m": 5 * 60_000,
    "15m": 15 * 60_000,
    "30m": 30 * 60_000,
    "1H": 60 * 60_000,
    "2H": 2 * 60 * 60_000,
    "4H": 4 * 60 * 60_000,
    "1D": 24 * 60 * 60_000,
}


def _coverage_exchange(sources: list[str] | None) -> tuple[str | None, bool]:
    """Collapse a canonical source_primary list into (label, mixed).

    Returns the source exchange(s) backing a coverage row. `mixed` is True when
    more than one source feeds the same symbol/bar (Binance-preferred fill with
    OKX gap-fill is the expected case, not corruption).
    """
    vals = sorted({str(s).lower() for s in (sources or []) if s})
    if not vals:
        return None, False
    return "+".join(vals), len(vals) > 1


class FetchRequest(BaseModel):
    symbol: str | None = None
    symbols: list[str] = Field(default_factory=list)
    exchange: str = "okx"
    existing_only: bool = False
    bar: str
    start: str
    end: str


class ExternalRefreshRequest(BaseModel):
    dataset_id: str | None = None
    dataset_ids: list[str] = Field(default_factory=list)
    start: str = ""
    end: str = ""


class _FetchCancelled(Exception):
    pass


def make_data_router(db_dsn: str | None = None) -> APIRouter:
    router = APIRouter()

    @router.get("/instruments")
    async def get_instruments(
        inst_type: str = "SWAP",
        quote_ccy: str = "USDT",
        exchange: str = "okx",
        q: str = "",
    ):
        exchange = _normalize_fetch_exchange(exchange)
        keyword = str(q or "").strip().upper()
        if exchange == "binance":
            rows = await _binance_instruments(quote_ccy=quote_ccy, keyword=keyword)
            return rows

        from okx_quant.data.exchange_clients.okx_public import OKXPublicClient

        client = OKXPublicClient()
        try:
            raw = await asyncio.to_thread(client.get_instruments, inst_type)
        finally:
            client.close()
        rows = []
        for item in raw:
            inst_id = item.get("instId") or ""
            inferred_quote = item.get("quoteCcy") or item.get("settleCcy")
            if quote_ccy and inferred_quote != quote_ccy and f"-{quote_ccy}-" not in inst_id:
                continue
            list_time_ms = _safe_int(item.get("listTime"))
            base_ccy = item.get("baseCcy") or inst_id.split("-")[0]
            row = {
                "exchange": "okx",
                "inst_id": inst_id,
                "native_symbol": inst_id,
                "normalized_symbol": inst_id,
                "inst_type": item.get("instType"),
                "base_ccy": base_ccy,
                "quote_ccy": inferred_quote,
                "settle_ccy": item.get("settleCcy"),
                "state": item.get("state"),
                "list_time_ms": list_time_ms,
                "list_date": _ms_to_date(list_time_ms),
            }
            if keyword and not _instrument_matches(row, keyword):
                continue
            rows.append(row)
        rows.sort(key=lambda r: _instrument_sort_key(r, keyword))
        return rows

    @router.get("/coverage")
    async def get_coverage():
        if not db_dsn:
            raise HTTPException(status_code=503, detail="DATABASE_URL not configured")
        import asyncpg

        conn = await asyncpg.connect(db_dsn)
        try:
            rows = await conn.fetch(
                """
                SELECT inst_id, bar,
                       MIN(ts) AS first_ts, MAX(ts) AS last_ts,
                       COUNT(*) AS row_count,
                       array_agg(DISTINCT source_primary) AS sources
                FROM canonical_candles
                GROUP BY inst_id, bar
                ORDER BY inst_id, bar
                """
            )
            fr = await conn.fetch(
                """
                SELECT inst_id, 'funding' AS bar,
                       MIN(ts) AS first_ts, MAX(ts) AS last_ts,
                       COUNT(*) AS row_count,
                       array_agg(DISTINCT source) AS sources
                FROM funding_rates
                GROUP BY inst_id
                ORDER BY inst_id
                """
            )
            external = await _fetch_external_coverage(conn)

            def _row(r: dict, *, data_kind: str, provider: str) -> dict:
                d = dict(r)
                exchange, mixed = _coverage_exchange(d.pop("sources", None))
                return {**d, "gap_count": 0, "data_kind": data_kind,
                        "provider": provider, "exchange": exchange, "mixed": mixed}

            return [
                _row(r, data_kind="ohlcv", provider="canonical") for r in rows
            ] + [
                _row(r, data_kind="funding", provider="okx") for r in fr
            ] + external
        finally:
            await conn.close()

    @router.get("/exchanges")
    async def get_exchanges():
        """List exchanges that have OHLCV data ingested into market_klines.

        Frontend uses this to populate the Run Backtest exchange dropdown.
        Falls back to a hardcoded default list when the DB is not configured,
        so the dropdown still works on parquet-only environments.
        """
        default = [
            {"exchange": "binance", "available": False, "row_count": 0},
            {"exchange": "okx", "available": False, "row_count": 0},
            {"exchange": "bybit", "available": False, "row_count": 0},
        ]
        if not db_dsn:
            return default
        import asyncpg
        try:
            conn = await asyncpg.connect(db_dsn)
        except Exception:
            return default
        try:
            rows = await conn.fetch(
                """
                SELECT exchange, COUNT(*)::bigint AS row_count
                FROM market_klines
                GROUP BY exchange
                ORDER BY exchange
                """
            )
            seen = {r["exchange"]: int(r["row_count"]) for r in rows}
            out = [
                {"exchange": ex, "available": ex in seen, "row_count": seen.get(ex, 0)}
                for ex in ("binance", "okx", "bybit", "coinbase", "kraken")
            ]
            # Surface any extra exchanges the DB knows about that aren't in our defaults.
            for extra in sorted(set(seen) - {row["exchange"] for row in out}):
                out.append({"exchange": extra, "available": True, "row_count": seen[extra]})
            return out
        except Exception:
            return default
        finally:
            await conn.close()

    @router.get("/export")
    async def export_data(
        symbols: str = "",
        bar: str = "1m",
        start: str = "",
        end: str = "",
        format: str = "xlsx",
        kind: str = "ohlcv",
        datasets: str = "",
    ):
        if not db_dsn:
            raise HTTPException(status_code=503, detail="DATABASE_URL not configured")
        data_kind = (kind or "ohlcv").strip().lower()
        if data_kind not in {"ohlcv", "funding", "external"}:
            raise HTTPException(status_code=400, detail="kind must be one of: ohlcv, funding, external")
        if not start or not end:
            raise HTTPException(status_code=400, detail="start and end are required")
        start_dt = _parse_utc(start)
        end_dt = _parse_utc(end)
        if start_dt >= end_dt:
            raise HTTPException(status_code=400, detail="start must be earlier than end")
        fmt = "xlsx" if format == "xlsx" else "csv"

        if data_kind == "external":
            dataset_ids = [s.strip() for s in datasets.split(",") if s.strip()]
            if not dataset_ids:
                raise HTTPException(status_code=400, detail="At least one dataset is required")
            filename = _export_filename(dataset_ids, "external", start_dt, end_dt, fmt=fmt)
            headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
            if fmt == "xlsx":
                content = await _build_external_xlsx(db_dsn, dataset_ids, start_dt, end_dt)
                return Response(
                    content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers,
                )
            return StreamingResponse(
                _stream_external_csv(db_dsn, dataset_ids, start_dt, end_dt),
                media_type="text/csv",
                headers=headers,
            )

        inst_ids = [s.strip() for s in symbols.split(",") if s.strip()]
        if not inst_ids:
            raise HTTPException(status_code=400, detail="At least one symbol is required")
        filename = _export_filename(inst_ids, bar if data_kind == "ohlcv" else "funding", start_dt, end_dt, fmt=fmt)
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        if data_kind == "funding":
            if fmt == "xlsx":
                content = await _build_funding_xlsx(db_dsn, inst_ids, start_dt, end_dt)
                return Response(
                    content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers,
                )
            return StreamingResponse(
                _stream_funding_csv(db_dsn, inst_ids, start_dt, end_dt),
                media_type="text/csv",
                headers=headers,
            )
        if fmt == "xlsx":
            content = await _build_ohlcv_xlsx(db_dsn, inst_ids, bar, start_dt, end_dt)
            return Response(
                content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers,
            )
        return StreamingResponse(
            _stream_ohlcv_csv(db_dsn, inst_ids, bar, start_dt, end_dt),
            media_type="text/csv",
            headers=headers,
        )

    @router.post("/external/refresh")
    async def refresh_external_data(req: ExternalRefreshRequest):
        if not db_dsn:
            raise HTTPException(status_code=503, detail="DATABASE_URL not configured")
        dataset_ids = _request_dataset_ids(req)
        if not dataset_ids:
            raise HTTPException(status_code=400, detail="At least one dataset is required")
        if not req.start or not req.end:
            raise HTTPException(status_code=400, detail="start and end are required")
        start_dt = _parse_utc(req.start)
        end_dt = _parse_utc(req.end)
        if start_dt >= end_dt:
            raise HTTPException(status_code=400, detail="start must be earlier than end")
        return await _refresh_external_datasets(db_dsn, dataset_ids, start_dt, end_dt)

    @router.post("/fetch")
    async def trigger_fetch(req: FetchRequest, bg: BackgroundTasks):
        if not db_dsn:
            raise HTTPException(status_code=503, detail="DATABASE_URL not configured")
        exchange = _normalize_fetch_exchange(req.exchange)
        symbols = await _resolve_fetch_symbols(req, db_dsn)
        if not symbols:
            detail = "No existing DB trading pairs found for this bar" if req.existing_only else "At least one symbol is required"
            raise HTTPException(status_code=400, detail=detail)
        job_id = str(uuid.uuid4())[:8]
        _jobs[job_id] = {
            "job_id": job_id,
            "exchange": exchange,
            "existing_only": bool(req.existing_only),
            "status": "queued",
            "progress": 0,
            "message": (
                f"Queued: {exchange.upper()} {len(symbols)} existing DB symbol update..."
                if req.existing_only
                else f"Queued: {exchange.upper()} {len(symbols)} symbol fetch..."
            ),
            "symbols": symbols,
            "symbol_count": len(symbols),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        bg.add_task(_run_fetch, job_id, req, db_dsn)
        return _jobs[job_id]

    @router.get("/fetch/status/{job_id}")
    async def fetch_status(job_id: str):
        if job_id not in _jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        return _jobs[job_id]

    @router.post("/fetch/cancel/{job_id}")
    async def cancel_fetch(job_id: str):
        if job_id not in _jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        job = _jobs[job_id]
        if job.get("status") in {"done", "error", "cancelled"}:
            return job
        job.update({
            "cancel_requested": True,
            "status": "cancelling",
            "message": f"Cancel requested for {job.get('exchange', '').upper() or 'data'} fetch...",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        return job

    @router.get("/fetch/jobs")
    async def fetch_jobs():
        return sorted(
            _jobs.values(),
            key=lambda row: str(row.get("updated_at") or row.get("created_at") or ""),
            reverse=True,
        )

    @router.delete("/pairs/{inst_id}")
    async def delete_pair(inst_id: str):
        if not db_dsn:
            raise HTTPException(status_code=503, detail="DATABASE_URL not configured")
        inst_id = str(inst_id or "").strip().upper()
        if not inst_id:
            raise HTTPException(status_code=400, detail="inst_id is required")
        if _active_job_for_symbol(inst_id):
            raise HTTPException(status_code=409, detail="Pair has an active fetch job; cancel it first")
        import asyncpg

        deleted: dict[str, int] = {}
        conn = await asyncpg.connect(db_dsn)
        try:
            async with conn.transaction():
                for sql, params in _pair_delete_statements(inst_id):
                    table = sql.split("FROM")[1].split()[0]
                    status = await conn.execute(sql, *params)
                    deleted[table] = int(status.rsplit(" ", 1)[-1]) if status.startswith("DELETE") else 0
        finally:
            await conn.close()
        parquet_removed, parquet_error = _remove_pair_parquet(
            inst_id,
            _project_root_path() / "data" / "ticks",
        )
        return {
            "inst_id": inst_id,
            "deleted": deleted,
            "parquet_removed": parquet_removed,
            "parquet_error": parquet_error,
        }

    return router


def _safe_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _job_cancel_requested(job_id: str) -> bool:
    return bool(_jobs.get(job_id, {}).get("cancel_requested"))


def _raise_if_fetch_cancelled(job_id: str) -> None:
    if _job_cancel_requested(job_id):
        raise _FetchCancelled()


def _mark_fetch_cancelled(job_id: str, message: str = "Fetch cancelled by user") -> None:
    job = _jobs.get(job_id)
    if not job:
        return
    job.update({
        "status": "cancelled",
        "cancel_requested": True,
        "message": message,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })


def _ms_to_date(value: int | None) -> str | None:
    if value is None:
        return None
    return datetime.fromtimestamp(value / 1000, tz=timezone.utc).date().isoformat()


def _normalize_fetch_exchange(value: str | None) -> str:
    exchange = str(value or "okx").strip().lower()
    if exchange not in {"okx", "binance"}:
        raise HTTPException(status_code=400, detail="exchange must be one of: okx, binance")
    return exchange


def _binance_native_to_normalized(symbol: str, quote: str = "USDT") -> str:
    clean = re.sub(r"[^A-Za-z0-9]", "", str(symbol or "").upper())
    quote = str(quote or "USDT").upper()
    if clean.endswith(quote) and len(clean) > len(quote):
        base = clean[: -len(quote)]
        return f"{base}-{quote}-SWAP"
    return clean


def _normalized_to_binance_native(symbol: str) -> str:
    parts = str(symbol or "").upper().split("-")
    if len(parts) >= 2:
        if parts[-1] in {"SWAP", "PERP", "FUTURES"}:
            parts = parts[:-1]
        return "".join(parts)
    return re.sub(r"[^A-Za-z0-9]", "", str(symbol or "").upper())


def _instrument_matches(row: dict[str, Any], keyword: str) -> bool:
    haystack = " ".join(
        str(row.get(key) or "")
        for key in ("inst_id", "native_symbol", "normalized_symbol", "base_ccy", "quote_ccy")
    ).upper()
    return keyword in haystack


def _instrument_sort_key(row: dict[str, Any], keyword: str = "") -> tuple[int, str]:
    inst_id = str(row.get("inst_id") or "").upper()
    native = str(row.get("native_symbol") or "").upper()
    normalized = str(row.get("normalized_symbol") or inst_id).upper()
    base = str(row.get("base_ccy") or "").upper()
    quote = str(row.get("quote_ccy") or "USDT").upper()
    keyword = str(keyword or "").upper()
    exact_normalized = f"{keyword}-{quote}-SWAP" if keyword and quote else ""
    exact_native = f"{keyword}{quote}" if keyword and quote else ""

    if keyword:
        if base == keyword or inst_id == exact_normalized or normalized == exact_normalized or native == exact_native:
            return (0, inst_id)
        if base.startswith(keyword) or inst_id.startswith(keyword) or native.startswith(keyword):
            return (1, inst_id)
        if keyword in inst_id or keyword in native or keyword in normalized:
            return (2, inst_id)
    return (3, inst_id)


def _positive_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    return out if math.isfinite(out) and out > 0 else None


def _binance_filter_value(item: dict[str, Any], filter_type: str, key: str) -> float | None:
    wanted = str(filter_type or "").upper()
    for filt in item.get("filters") or []:
        if str(filt.get("filterType") or "").upper() == wanted:
            return _positive_float(filt.get(key))
    return None


def _binance_venue_spec_from_meta(meta: dict[str, Any]) -> dict[str, float | str] | None:
    tick_size = _positive_float(meta.get("tick_size"))
    lot_size = _positive_float(meta.get("lot_size"))
    min_size = _positive_float(meta.get("min_size"))
    if tick_size is None or lot_size is None or min_size is None:
        return None
    return {
        "ct_val": 1.0,
        "lot_size": lot_size,
        "tick_size": tick_size,
        "min_size": min_size,
        "source": "binance_exchange_info",
    }


async def _upsert_venue_instrument_spec(
    pool: Any,
    *,
    exchange: str,
    symbol: str,
    ct_val: float,
    lot_size: float,
    tick_size: float,
    min_size: float,
    source: str,
) -> None:
    await pool.execute(
        """
        INSERT INTO venue_instrument_specs
            (exchange, symbol, ct_val, lot_size, tick_size, min_size, source)
        VALUES ($1, $2, $3, $4, $5, $6, $7)
        ON CONFLICT (exchange, symbol) DO UPDATE SET
            ct_val = EXCLUDED.ct_val,
            lot_size = EXCLUDED.lot_size,
            tick_size = EXCLUDED.tick_size,
            min_size = EXCLUDED.min_size,
            source = EXCLUDED.source,
            updated_at = NOW()
        """,
        exchange,
        symbol,
        float(ct_val),
        float(lot_size),
        float(tick_size),
        float(min_size),
        source,
    )


async def _binance_instruments(quote_ccy: str = "USDT", keyword: str = "") -> list[dict[str, Any]]:
    from okx_quant.data.exchange_clients.binance_public import BinancePublicClient

    quote = str(quote_ccy or "USDT").upper()
    client = BinancePublicClient()
    try:
        payload = await asyncio.to_thread(client.get_futures_exchange_info)
    finally:
        client.close()
    rows: list[dict[str, Any]] = []
    for item in payload.get("symbols") or []:
        native = str(item.get("symbol") or "").upper()
        base = str(item.get("baseAsset") or "").upper()
        quote_asset = str(item.get("quoteAsset") or "").upper()
        margin = str(item.get("marginAsset") or quote_asset or quote).upper()
        if quote and quote_asset != quote:
            continue
        if item.get("contractType") not in {None, "PERPETUAL"}:
            continue
        normalized = _binance_native_to_normalized(native, quote_asset or quote)
        list_time_ms = _safe_int(item.get("onboardDate"))
        row = {
            "exchange": "binance",
            "inst_id": normalized,
            "native_symbol": native,
            "normalized_symbol": normalized,
            "inst_type": "SWAP",
            "base_ccy": base or normalized.split("-")[0],
            "quote_ccy": quote_asset or quote,
            "settle_ccy": margin,
            "state": item.get("status"),
            "list_time_ms": list_time_ms,
            "list_date": _ms_to_date(list_time_ms),
            "tick_size": _binance_filter_value(item, "PRICE_FILTER", "tickSize"),
            "lot_size": _binance_filter_value(item, "LOT_SIZE", "stepSize"),
            "min_size": _binance_filter_value(item, "LOT_SIZE", "minQty"),
        }
        if keyword and not _instrument_matches(row, keyword):
            continue
        rows.append(row)
    rows.sort(key=lambda r: _instrument_sort_key(r, keyword))
    return rows


def _okx_instrument_rows(
    raw: list[dict[str, Any]],
    quote_ccy: str = "USDT",
    keyword: str = "",
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in raw:
        inst_id = item.get("instId") or ""
        inferred_quote = item.get("quoteCcy") or item.get("settleCcy")
        if quote_ccy and inferred_quote != quote_ccy and f"-{quote_ccy}-" not in inst_id:
            continue
        list_time_ms = _safe_int(item.get("listTime"))
        base_ccy = item.get("baseCcy") or inst_id.split("-")[0]
        row = {
            "exchange": "okx",
            "inst_id": inst_id,
            "native_symbol": inst_id,
            "normalized_symbol": inst_id,
            "inst_type": item.get("instType"),
            "base_ccy": base_ccy,
            "quote_ccy": inferred_quote,
            "settle_ccy": item.get("settleCcy"),
            "state": item.get("state"),
            "list_time_ms": list_time_ms,
            "list_date": _ms_to_date(list_time_ms),
        }
        if keyword and not _instrument_matches(row, keyword):
            continue
        rows.append(row)
    rows.sort(key=lambda r: _instrument_sort_key(r, keyword))
    return rows


def _parse_utc(value: str) -> datetime:
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid datetime: {value}") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _export_filename(inst_ids: list[str], bar: str, start: datetime, end: datetime, fmt: str = "xlsx") -> str:
    bar = _normalize_output_bar(bar)
    ext = "xlsx" if fmt == "xlsx" else "csv"
    prefix = "ohlcv_export"
    if bar == "funding":
        prefix = "funding_export"
    elif bar == "external":
        prefix = "external_export"
    return f"{prefix}_{bar}_{start.date()}_{end.date()}.{ext}"


def _is_hourly_aggregate(bar: str) -> bool:
    return bar.lower() in {"1h", "1hr", "1hour"}


def _normalize_output_bar(bar: str) -> str:
    return "1H" if _is_hourly_aggregate(bar) else bar


def _export_select_sql(hourly: bool) -> str:
    if hourly:
        return """
            WITH hourly AS (
                SELECT
                    date_trunc('hour', ts) AS ts,
                    inst_id,
                    (array_agg(open ORDER BY ts))[1] AS open,
                    MAX(high) AS high,
                    MIN(low) AS low,
                    (array_agg(close ORDER BY ts DESC))[1] AS close,
                    SUM(vol_contract) AS vol_contract,
                    SUM(vol_base) AS vol_base,
                    SUM(vol_quote) AS vol_quote,
                    CASE
                        WHEN COUNT(DISTINCT source_primary) = 1 THEN MIN(source_primary)
                        ELSE 'mixed'
                    END AS source_primary,
                    CASE
                        WHEN BOOL_OR(quality_status = 'suspect') THEN 'suspect'
                        ELSE 'raw'
                    END AS quality_status
                FROM canonical_candles
                WHERE inst_id = ANY($1::text[])
                  AND bar = '1m'
                  AND ts >= $3
                  AND ts < $4
                GROUP BY inst_id, date_trunc('hour', ts)
            )
            SELECT
                ts,
                inst_id,
                $2::text AS bar,
                open,
                high,
                low,
                close,
                vol_contract,
                vol_base,
                vol_quote,
                source_primary,
                quality_status
            FROM hourly
            ORDER BY inst_id, ts
        """
    return """
        SELECT
            ts,
            inst_id,
            bar,
            open,
            high,
            low,
            close,
            vol_contract,
            vol_base,
            vol_quote,
            source_primary,
            quality_status
        FROM canonical_candles
        WHERE inst_id = ANY($1::text[])
          AND bar = $2
          AND ts >= $3
          AND ts < $4
        ORDER BY inst_id, ts
    """


async def _stream_ohlcv_csv(
    db_dsn: str,
    inst_ids: list[str],
    bar: str,
    start: datetime,
    end: datetime,
):
    import asyncpg

    conn = await asyncpg.connect(db_dsn)
    output = io.StringIO()
    writer = csv.writer(output)
    output_bar = _normalize_output_bar(bar)
    hourly = _is_hourly_aggregate(bar)
    columns = [
        "ts",
        "inst_id",
        "bar",
        "open",
        "high",
        "low",
        "close",
        "vol_contract",
        "vol_base",
        "vol_quote",
        "source_primary",
        "quality_status",
    ]
    try:
        writer.writerow(columns)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        async with conn.transaction():
            pending = 0
            async for row in conn.cursor(
                _export_select_sql(hourly),
                inst_ids,
                output_bar,
                start,
                end,
                prefetch=10_000,
            ):
                writer.writerow([
                    row["ts"].isoformat(),
                    row["inst_id"],
                    row["bar"],
                    row["open"],
                    row["high"],
                    row["low"],
                    row["close"],
                    row["vol_contract"],
                    row["vol_base"],
                    row["vol_quote"],
                    row["source_primary"],
                    row["quality_status"],
                ])
                pending += 1
                if pending >= 10_000:
                    yield output.getvalue()
                    output.seek(0)
                    output.truncate(0)
                    pending = 0
            if pending:
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)
    finally:
        await conn.close()


async def _build_ohlcv_xlsx(
    db_dsn: str,
    inst_ids: list[str],
    bar: str,
    start: datetime,
    end: datetime,
) -> bytes:
    import asyncpg
    import openpyxl

    conn = await asyncpg.connect(db_dsn)
    output_bar = _normalize_output_bar(bar)
    hourly = _is_hourly_aggregate(bar)
    columns = [
        "ts", "inst_id", "bar", "open", "high", "low", "close",
        "vol_contract", "vol_base", "vol_quote", "source_primary", "quality_status",
    ]

    wb = openpyxl.Workbook(write_only=True)
    sheets: dict[str, Any] = {}
    for inst_id in inst_ids:
        ws = wb.create_sheet(title=inst_id[:31])
        ws.append(columns)
        sheets[inst_id] = ws

    try:
        async with conn.transaction():
            async for row in conn.cursor(
                _export_select_sql(hourly),
                inst_ids,
                output_bar,
                start,
                end,
                prefetch=10_000,
            ):
                sheets[row["inst_id"]].append([
                    row["ts"].isoformat(),
                    row["inst_id"],
                    row["bar"],
                    row["open"],
                    row["high"],
                    row["low"],
                    row["close"],
                    row["vol_contract"],
                    row["vol_base"],
                    row["vol_quote"],
                    row["source_primary"],
                    row["quality_status"],
                ])
    finally:
        await conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FUNDING_EXPORT_COLUMNS = [
    "ts",
    "source",
    "inst_id",
    "funding_rate",
    "realized_rate",
    "mark_price",
    "funding_interval_hours",
    "next_funding_ts",
    "apr",
]


EXTERNAL_EXPORT_COLUMNS = [
    "dataset_id",
    "observed_at",
    "published_at",
    "value_num",
    "value_text",
    "open",
    "high",
    "low",
    "close",
    "volume",
    "ticker",
    "interval",
    "quality_status",
    "provider",
    "attribution",
    "research_only",
    "source_caveat",
]


def _funding_select_sql() -> str:
    return """
        SELECT
            ts,
            source,
            inst_id,
            funding_rate,
            realized_rate,
            mark_price,
            funding_interval_hours,
            next_funding_ts,
            funding_rate * (365 * 24 / COALESCE(NULLIF(funding_interval_hours, 0), 8.0)) AS apr
        FROM funding_rates
        WHERE inst_id = ANY($1::text[])
          AND ts >= $2
          AND ts < $3
        ORDER BY inst_id, ts
    """


async def _stream_funding_csv(
    db_dsn: str,
    inst_ids: list[str],
    start: datetime,
    end: datetime,
):
    import asyncpg

    conn = await asyncpg.connect(db_dsn)
    output = io.StringIO()
    writer = csv.writer(output)
    try:
        writer.writerow(FUNDING_EXPORT_COLUMNS)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        async with conn.transaction():
            pending = 0
            async for row in conn.cursor(_funding_select_sql(), inst_ids, start, end, prefetch=10_000):
                writer.writerow(_funding_export_row(row))
                pending += 1
                if pending >= 10_000:
                    yield output.getvalue()
                    output.seek(0)
                    output.truncate(0)
                    pending = 0
            if pending:
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)
    finally:
        await conn.close()


async def _build_funding_xlsx(
    db_dsn: str,
    inst_ids: list[str],
    start: datetime,
    end: datetime,
) -> bytes:
    import asyncpg
    import openpyxl

    conn = await asyncpg.connect(db_dsn)
    wb = openpyxl.Workbook(write_only=True)
    sheets: dict[str, Any] = {}
    for inst_id in inst_ids:
        ws = wb.create_sheet(title=_sheet_title(inst_id))
        ws.append(FUNDING_EXPORT_COLUMNS)
        sheets[inst_id] = ws
    try:
        async with conn.transaction():
            async for row in conn.cursor(_funding_select_sql(), inst_ids, start, end, prefetch=10_000):
                sheets[row["inst_id"]].append(_funding_export_row(row))
    finally:
        await conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _funding_export_row(row: Any) -> list[Any]:
    return [
        row["ts"].isoformat(),
        row["source"],
        row["inst_id"],
        row["funding_rate"],
        row["realized_rate"],
        row["mark_price"],
        row["funding_interval_hours"],
        row["next_funding_ts"].isoformat() if row["next_funding_ts"] else None,
        row["apr"],
    ]


def _external_select_sql() -> str:
    return """
        SELECT
            o.dataset_id,
            o.observed_at,
            o.published_at,
            o.value_num,
            o.value_text,
            o.fields,
            o.quality_status,
            d.provider,
            d.attribution,
            COALESCE((d.metadata->>'research_only')::boolean, false) AS research_only
        FROM external_observations o
        JOIN external_datasets d ON d.dataset_id = o.dataset_id
        WHERE o.dataset_id = ANY($1::text[])
          AND o.observed_at >= $2
          AND o.observed_at < $3
        ORDER BY o.dataset_id, o.observed_at
    """


async def _stream_external_csv(
    db_dsn: str,
    dataset_ids: list[str],
    start: datetime,
    end: datetime,
):
    import asyncpg

    conn = await asyncpg.connect(db_dsn)
    output = io.StringIO()
    writer = csv.writer(output)
    try:
        writer.writerow(EXTERNAL_EXPORT_COLUMNS)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        async with conn.transaction():
            pending = 0
            async for row in conn.cursor(_external_select_sql(), dataset_ids, start, end, prefetch=10_000):
                writer.writerow(_external_export_row(row))
                pending += 1
                if pending >= 10_000:
                    yield output.getvalue()
                    output.seek(0)
                    output.truncate(0)
                    pending = 0
            if pending:
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)
    finally:
        await conn.close()


async def _build_external_xlsx(
    db_dsn: str,
    dataset_ids: list[str],
    start: datetime,
    end: datetime,
) -> bytes:
    import asyncpg
    import openpyxl

    conn = await asyncpg.connect(db_dsn)
    wb = openpyxl.Workbook(write_only=True)
    sheets: dict[str, Any] = {}
    for dataset_id in dataset_ids:
        ws = wb.create_sheet(title=_sheet_title(dataset_id))
        ws.append(EXTERNAL_EXPORT_COLUMNS)
        sheets[dataset_id] = ws
    try:
        async with conn.transaction():
            async for row in conn.cursor(_external_select_sql(), dataset_ids, start, end, prefetch=10_000):
                sheets[row["dataset_id"]].append(_external_export_row(row))
    finally:
        await conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _external_export_row(row: Any) -> list[Any]:
    fields = row["fields"] or {}
    if isinstance(fields, str):
        import json as _json
        try:
            fields = _json.loads(fields)
        except Exception:
            fields = {}
    return [
        row["dataset_id"],
        row["observed_at"].isoformat(),
        row["published_at"].isoformat() if row["published_at"] else None,
        row["value_num"],
        row["value_text"],
        fields.get("open"),
        fields.get("high"),
        fields.get("low"),
        fields.get("close"),
        fields.get("volume"),
        fields.get("ticker"),
        fields.get("interval"),
        row["quality_status"],
        row["provider"],
        row["attribution"],
        row["research_only"],
        fields.get("source_caveat"),
    ]


def _sheet_title(value: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", value)[:31] or "sheet"


async def _refresh_external_datasets(
    db_dsn: str,
    dataset_ids: list[str],
    start: datetime,
    end: datetime,
) -> dict[str, Any]:
    from okx_quant.data.external_clients.yfinance_client import YFinanceClient
    from okx_quant.data.external_store import ExternalDataStore
    import yaml

    config_path = _project_root_path() / "config" / "external_data.yaml"
    with config_path.open("r", encoding="utf-8") as fh:
        datasets = (yaml.safe_load(fh) or {}).get("datasets") or {}

    refreshed: list[dict[str, Any]] = []
    async with await ExternalDataStore.from_dsn(db_dsn, min_size=1, max_size=2) as store:
        for dataset_id in dataset_ids:
            cfg = datasets.get(dataset_id)
            if not cfg:
                raise HTTPException(status_code=400, detail=f"Unknown external dataset: {dataset_id}")
            if str(cfg.get("adapter") or "") != "yfinance":
                raise HTTPException(status_code=400, detail=f"Dataset does not support on-demand refresh: {dataset_id}")
            await store.upsert_dataset(dataset_id, cfg)
            job_id = await store.start_fetch_job(dataset_id, str(cfg.get("provider") or "yfinance"), start, end)
            try:
                client = YFinanceClient(publish_lag_days=int(cfg.get("publish_lag_days", 1)))
                rows = await asyncio.to_thread(
                    client.fetch,
                    ticker=str(cfg.get("ticker") or "BTC=F"),
                    start=start,
                    end=end,
                    interval=str(cfg.get("interval") or "1d"),
                )
                if not rows and bool(cfg.get("fail_on_empty_fetch", False)):
                    raise RuntimeError(f"{dataset_id}: empty yfinance fetch")
                stats = await store.upsert_observations(dataset_id, rows)
                await store.finish_fetch_job(
                    job_id,
                    status="success",
                    rows_fetched=len(rows),
                    rows_inserted=stats["inserted"],
                    rows_updated=stats["updated"],
                    details={"on_demand_export_refresh": True},
                )
                await store.update_checkpoint(
                    dataset_id,
                    direction="backfill",
                    cursor_time=max((row["observed_at"] for row in rows), default=end),
                    request_count=1,
                    row_count=len(rows),
                    status="success",
                )
                refreshed.append({
                    "dataset_id": dataset_id,
                    "status": "success",
                    "rows_fetched": len(rows),
                    "rows_inserted": stats["inserted"],
                    "rows_updated": stats["updated"],
                    "research_only": bool(cfg.get("research_only", False)),
                    "attribution": cfg.get("attribution"),
                })
            except Exception as exc:
                await store.finish_fetch_job(job_id, status="failed", error_message=str(exc))
                await store.update_checkpoint(
                    dataset_id,
                    direction="backfill",
                    cursor_time=start,
                    request_count=1,
                    row_count=0,
                    status="failed",
                    last_error=str(exc),
                )
                raise HTTPException(status_code=502, detail=str(exc)) from exc
    return {"status": "done", "datasets": refreshed}


def _project_root_path():
    from pathlib import Path as _Path

    return _Path(__file__).resolve().parents[3]


def _pair_delete_statements(inst_id: str) -> list[tuple[str, list]]:
    subquery = (
        "instrument_id IN (SELECT instrument_id FROM market_instruments "
        "WHERE canonical_inst_id=$1 OR normalized_symbol=$1)"
    )
    return [
        (f"DELETE FROM market_klines WHERE {subquery}", [inst_id]),
        (f"DELETE FROM market_funding_rates WHERE {subquery}", [inst_id]),
        ("DELETE FROM market_instruments WHERE canonical_inst_id=$1 OR normalized_symbol=$1", [inst_id]),
        ("DELETE FROM canonical_candles WHERE inst_id=$1", [inst_id]),
        ("DELETE FROM raw_candles WHERE inst_id=$1", [inst_id]),
        ("DELETE FROM funding_rates WHERE inst_id=$1", [inst_id]),
        ("DELETE FROM instrument_bars WHERE inst_id=$1", [inst_id]),
        ("DELETE FROM instruments WHERE inst_id=$1", [inst_id]),
    ]


def _active_job_for_symbol(inst_id: str) -> bool:
    target = str(inst_id or "").upper()
    for job in _jobs.values():
        if job.get("status") in _TERMINAL_FETCH_STATUSES:
            continue
        symbols = [str(symbol or "").upper() for symbol in (job.get("symbols") or [])]
        if target in symbols:
            return True
    return False


def _remove_pair_parquet(inst_id: str, ticks_dir: str | Path) -> tuple[bool, str | None]:
    inst_dir = Path(ticks_dir) / str(inst_id).replace("-", "_")
    if not inst_dir.exists():
        return False, None
    try:
        shutil.rmtree(inst_dir)
        return True, None
    except Exception as exc:
        return False, str(exc)


async def _write_fetched_to_parquet(
    db_dsn: str,
    inst_id: str,
    bar: str,
    start_dt: datetime,
    end_dt: datetime,
) -> int:
    """
    Read freshly-canonicalized candles from DB and upsert into the local parquet file.
    Returns the number of rows written.  Non-fatal — caller should catch exceptions.
    """
    import sys
    from pathlib import Path as _Path

    import asyncpg
    import pandas as _pd

    _project_root = _Path(__file__).resolve().parents[3]
    _bt_path = str(_project_root / "backtesting")
    if _bt_path not in sys.path:
        sys.path.insert(0, _bt_path)
    from data_loader import write_candles_parquet  # type: ignore[import]

    conn = await asyncpg.connect(db_dsn)
    try:
        rows = await conn.fetch(
            """
            SELECT ts,
                   open, high, low, close,
                   COALESCE(vol_quote, vol_base, vol_contract, 0.0) AS vol
            FROM canonical_candles
            WHERE inst_id = $1 AND bar = $2
              AND ts >= $3 AND ts < $4
            ORDER BY ts
            """,
            inst_id, bar, start_dt, end_dt,
        )
    finally:
        await conn.close()

    if not rows:
        return 0

    df = _pd.DataFrame([dict(r) for r in rows])
    df["ts"] = _pd.to_datetime(df["ts"], utc=True).dt.tz_localize(None)
    df = df.set_index("ts")

    data_dir = str(_project_root / "data" / "ticks")
    write_candles_parquet(inst_id, bar, df, data_dir)
    return len(rows)


def _request_symbols(req: FetchRequest) -> list[str]:
    exchange = _normalize_fetch_exchange(req.exchange)
    symbols = list(req.symbols or [])
    if req.symbol:
        symbols.append(req.symbol)
    seen = set()
    normalized = []
    for raw in symbols:
        symbol = str(raw or "").strip().upper()
        if not symbol:
            continue
        if exchange == "binance":
            symbol = symbol if "-" in symbol else _binance_native_to_normalized(symbol)
        if symbol in seen:
            continue
        seen.add(symbol)
        normalized.append(symbol)
    return normalized


async def _existing_db_symbol_bounds(db_dsn: str, bar: str) -> list[dict[str, Any]]:
    import asyncpg

    conn = await asyncpg.connect(db_dsn)
    try:
        rows = await conn.fetch(
            """
            SELECT inst_id, MIN(ts) AS first_ts, MAX(ts) AS last_ts
            FROM canonical_candles
            WHERE bar = $1
            GROUP BY inst_id
            ORDER BY last_ts DESC, inst_id
            """,
            bar,
        )
    finally:
        await conn.close()
    return [
        {"inst_id": str(row["inst_id"]), "first_ts": row["first_ts"], "last_ts": row["last_ts"]}
        for row in rows
        if row["inst_id"]
    ]


async def _existing_db_symbols(db_dsn: str, bar: str) -> list[str]:
    rows = await _existing_db_symbol_bounds(db_dsn, bar)
    return [str(row["inst_id"]) for row in rows if row.get("inst_id")]


async def _resolve_fetch_symbols(req: FetchRequest, db_dsn: str) -> list[str]:
    if req.existing_only:
        return await _existing_db_symbols(db_dsn, req.bar)
    return _request_symbols(req)


def _coerce_utc_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _next_existing_update_start(
    requested_start_dt: datetime,
    existing_last_ts: Any,
    bar: str,
) -> datetime:
    last_dt = _coerce_utc_datetime(existing_last_ts)
    if not last_dt:
        return requested_start_dt
    next_dt = last_dt + timedelta(milliseconds=_FETCH_BAR_MS.get(bar, 60_000))
    return max(requested_start_dt, next_dt)


def _existing_update_ranges(
    requested_start_dt: datetime,
    requested_end_dt: datetime,
    existing_bounds: dict[str, Any] | None,
    bar: str,
) -> list[tuple[datetime, datetime]]:
    if requested_start_dt >= requested_end_dt:
        return []
    if not existing_bounds:
        return [(requested_start_dt, requested_end_dt)]

    first_dt = _coerce_utc_datetime(existing_bounds.get("first_ts"))
    last_dt = _coerce_utc_datetime(existing_bounds.get("last_ts"))
    if not first_dt or not last_dt:
        return [(requested_start_dt, requested_end_dt)]

    bar_delta = timedelta(milliseconds=_FETCH_BAR_MS.get(bar, 60_000))
    ranges: list[tuple[datetime, datetime]] = []
    if requested_start_dt < first_dt:
        ranges.append((requested_start_dt, min(first_dt, requested_end_dt)))

    next_dt = last_dt + bar_delta
    if next_dt < requested_end_dt:
        ranges.append((max(requested_start_dt, next_dt), requested_end_dt))

    return [(start, end) for start, end in ranges if start < end]


def _request_dataset_ids(req: ExternalRefreshRequest) -> list[str]:
    dataset_ids = list(req.dataset_ids or [])
    if req.dataset_id:
        dataset_ids.append(req.dataset_id)
    seen = set()
    return [s for s in dataset_ids if s and not (s in seen or seen.add(s))]


async def _fetch_external_coverage(conn: Any) -> list[dict[str, Any]]:
    try:
        rows = await conn.fetch(
            """
            SELECT
                d.dataset_id AS inst_id,
                COALESCE(d.frequency, 'external') AS bar,
                MIN(o.observed_at) AS first_ts,
                MAX(o.observed_at) AS last_ts,
                COUNT(o.observed_at) AS row_count,
                d.provider,
                d.value_kind,
                d.frequency,
                d.source_url,
                d.attribution,
                COALESCE((d.metadata->>'research_only')::boolean, false) AS research_only
            FROM external_datasets d
            LEFT JOIN external_observations o
              ON o.dataset_id = d.dataset_id
            GROUP BY
                d.dataset_id, d.provider, d.value_kind, d.frequency,
                d.source_url, d.attribution, d.metadata
            ORDER BY d.dataset_id
            """
        )
    except Exception:
        return []
    return [
        {
            **dict(row),
            "gap_count": 0,
            "data_kind": "external",
        }
        for row in rows
    ]


def _instrument_map(raw: list[dict]) -> dict[str, dict]:
    return {r.get("instId"): r for r in raw if r.get("instId")}


async def _run_fetch(job_id: str, req: FetchRequest, db_dsn: str) -> None:
    if _job_cancel_requested(job_id):
        _mark_fetch_cancelled(job_id)
        return
    job = _jobs.get(job_id)
    if job is not None:
        job.update({
            "status": "queued",
            "message": "Queued - waiting for running fetch...",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    async with _fetch_lock:
        if _job_cancel_requested(job_id):
            _mark_fetch_cancelled(job_id)
            return
        if job is not None:
            job.update({
                "status": "running",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
        await _run_fetch_body(job_id, req, db_dsn)


async def _run_fetch_body(job_id: str, req: FetchRequest, db_dsn: str) -> None:
    try:
        import asyncpg

        from okx_quant.data.candle_store import CandleStore
        from okx_quant.data.exchange_clients.binance_public import BinancePublicClient
        from okx_quant.data.exchange_clients.okx_public import OKXPublicClient

        exchange = _normalize_fetch_exchange(req.exchange)
        existing_bounds: dict[str, dict[str, Any]] = {}
        if req.existing_only:
            bound_rows = await _existing_db_symbol_bounds(db_dsn, req.bar)
            existing_bounds = {str(row["inst_id"]): row for row in bound_rows}
            symbols = list(existing_bounds)
        else:
            symbols = _request_symbols(req)
        requested_start_dt = datetime.fromisoformat(req.start).replace(tzinfo=timezone.utc)
        end_dt = datetime.fromisoformat(req.end).replace(tzinfo=timezone.utc)

        _jobs[job_id].update({
            "exchange": exchange,
            "existing_only": bool(req.existing_only),
            "symbols": symbols,
            "symbol_count": len(symbols),
            "created_at": _jobs[job_id].get("created_at") or datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "message": f"Loading {exchange.upper()} instrument metadata...",
        })
        _raise_if_fetch_cancelled(job_id)
        client = BinancePublicClient() if exchange == "binance" else OKXPublicClient()
        pool = await asyncpg.create_pool(db_dsn, min_size=1, max_size=3)
        try:
            if exchange == "binance":
                instruments = {
                    row["inst_id"]: row
                    for row in await _binance_instruments(quote_ccy="USDT")
                }
            else:
                raw = await asyncio.to_thread(client.get_instruments, "SWAP")
                instruments = {
                    row["inst_id"]: row
                    for row in _okx_instrument_rows(raw, quote_ccy="USDT", keyword="")
                }
            fetched: list[dict] = []
            total = len(symbols)
            store = CandleStore(pool)
            for idx, symbol in enumerate(symbols, start=1):
                _raise_if_fetch_cancelled(job_id)
                meta = instruments.get(symbol, {})
                list_time_ms = _safe_int(meta.get("listTime"))
                if list_time_ms is None:
                    list_time_ms = _safe_int(meta.get("list_time_ms"))
                list_dt = (
                    datetime.fromtimestamp(list_time_ms / 1000, tz=timezone.utc)
                    if list_time_ms else requested_start_dt
                )
                symbol_ranges = (
                    _existing_update_ranges(requested_start_dt, end_dt, existing_bounds.get(symbol), req.bar)
                    if req.existing_only
                    else [(requested_start_dt, end_dt)]
                )
                fetch_ranges = []
                for range_start_dt, range_end_dt in symbol_ranges:
                    start_dt = max(range_start_dt, list_dt)
                    if start_dt < range_end_dt:
                        fetch_ranges.append((start_dt, range_end_dt))
                symbol_progress_base = int(((idx - 1) / total) * 90)
                _jobs[job_id].update({
                    "progress": max(3, symbol_progress_base),
                    "message": f"Fetching {symbol} {req.bar} from {exchange.upper()} ({idx}/{total})...",
                    "current_symbol": symbol,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                })
                if exchange == "binance":
                    venue_spec = _binance_venue_spec_from_meta(meta)
                    if venue_spec is not None:
                        await _upsert_venue_instrument_spec(
                            pool,
                            exchange=exchange,
                            symbol=symbol,
                            ct_val=float(venue_spec["ct_val"]),
                            lot_size=float(venue_spec["lot_size"]),
                            tick_size=float(venue_spec["tick_size"]),
                            min_size=float(venue_spec["min_size"]),
                            source=str(venue_spec["source"]),
                        )
                    _raise_if_fetch_cancelled(job_id)
                if not fetch_ranges:
                    skip_start_dt = max(requested_start_dt, list_dt)
                    fetched.append({
                        "symbol": symbol,
                        "status": "skipped",
                        "exchange": exchange,
                        "rows": 0,
                        "list_date": _ms_to_date(list_time_ms),
                        "effective_start": skip_start_dt.date().isoformat(),
                        "effective_end": end_dt.date().isoformat(),
                        "message": (
                            "Already covered for requested date range"
                            if req.existing_only and existing_bounds.get(symbol)
                            else "Listing date is after requested end date"
                        ),
                    })
                    _jobs[job_id].update({
                        "progress": int((idx / total) * 90),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    })
                    continue

                native_symbol = str(meta.get("native_symbol") or symbol)
                if exchange == "binance":
                    native_symbol = _normalized_to_binance_native(symbol)
                base_ccy = meta.get("base_ccy") or symbol.split("-")[0]
                quote_ccy = meta.get("quote_ccy") or "USDT"
                settle_ccy = meta.get("settle_ccy") or quote_ccy
                await store.register_instrument(
                    inst_id=symbol,
                    base_ccy=base_ccy,
                    quote_ccy=quote_ccy,
                    settle_ccy=settle_ccy,
                    exchange=exchange,
                )
                _raise_if_fetch_cancelled(job_id)
                await store.register_instrument_bar(inst_id=symbol, bar=req.bar)
                instrument_id = await store.register_market_instrument(
                    exchange=exchange,
                    inst_id=native_symbol,
                    normalized_symbol=symbol,
                    base_asset=base_ccy,
                    quote_asset=quote_ccy,
                    settlement_asset=settle_ccy,
                    listing_time=list_dt if list_time_ms else None,
                    is_active=str(meta.get("state") or "").upper() not in {"CLOSE", "BREAK", "END_TRADING"},
                    canonical_inst_id=symbol,
                )

                rows_fetched = 0
                parquet_rows = 0
                parquet_errors: list[str] = []
                fetched_ranges = []
                range_total = len(fetch_ranges)
                for range_idx, (start_dt, range_end_dt) in enumerate(fetch_ranges, start=1):
                    start_ms = int(start_dt.timestamp() * 1000)
                    range_end_ms = int(range_end_dt.timestamp() * 1000)
                    range_label = f" range {range_idx}/{range_total}" if range_total > 1 else ""
                    _jobs[job_id].update({
                        "progress": max(3, symbol_progress_base),
                        "message": f"Fetching {symbol} {req.bar}{range_label} from {exchange.upper()} ({idx}/{total})...",
                        "current_symbol": symbol,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    })
                    if exchange == "binance":
                        candles = await asyncio.to_thread(
                            client.get_klines_range,
                            native_symbol,
                            req.bar,
                            start_ms,
                            range_end_ms,
                            should_cancel=lambda: _job_cancel_requested(job_id),
                        )
                    else:
                        candles = await asyncio.to_thread(
                            client.paginate_history,
                            native_symbol,
                            req.bar,
                            start_ms,
                            range_end_ms,
                            should_cancel=lambda: _job_cancel_requested(job_id),
                        )
                    _raise_if_fetch_cancelled(job_id)

                    _jobs[job_id].update({
                        "progress": min(95, int(((idx - 0.45) / total) * 90)),
                        "message": f"Writing {symbol} {req.bar}{range_label} to DB ({idx}/{total})...",
                        "current_symbol": symbol,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    })
                    await store.upsert_raw_candles(candles, source=exchange, inst_id=symbol, bar=req.bar)
                    _raise_if_fetch_cancelled(job_id)
                    await store.canonicalize_from_raw(
                        source=exchange,
                        inst_id=symbol,
                        bar=req.bar,
                        start=start_dt,
                        end=range_end_dt,
                    )
                    _raise_if_fetch_cancelled(job_id)
                    await store.upsert_market_klines(
                        candles,
                        instrument_id=instrument_id,
                        bar=req.bar,
                        data_source=exchange,
                    )
                    _raise_if_fetch_cancelled(job_id)
                    _jobs[job_id].update({
                        "progress": min(98, int((idx / total) * 90)),
                        "message": f"Exporting {symbol} {req.bar}{range_label} parquet ({idx}/{total})...",
                        "current_symbol": symbol,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    })

                    range_parquet_rows = 0
                    try:
                        range_parquet_rows = await _write_fetched_to_parquet(
                            db_dsn, symbol, req.bar, start_dt, range_end_dt
                        )
                    except Exception as exc:
                        parquet_errors.append(str(exc))
                    _raise_if_fetch_cancelled(job_id)
                    rows_fetched += len(candles)
                    parquet_rows += range_parquet_rows
                    fetched_ranges.append(
                        {
                            "start": start_dt.date().isoformat(),
                            "end": range_end_dt.date().isoformat(),
                            "rows": len(candles),
                        }
                    )
                await store.update_instrument_bar_bounds(symbol, req.bar)
                _raise_if_fetch_cancelled(job_id)

                fetched.append({
                    "symbol": symbol,
                    "native_symbol": native_symbol,
                    "exchange": exchange,
                    "status": "done",
                    "rows": rows_fetched,
                    "parquet_rows": parquet_rows,
                    "parquet_error": "; ".join(parquet_errors) if parquet_errors else None,
                    "list_date": _ms_to_date(list_time_ms),
                    "effective_start": fetch_ranges[0][0].date().isoformat(),
                    "effective_end": fetch_ranges[-1][1].date().isoformat(),
                    "ranges": fetched_ranges,
                })
        finally:
            client.close()
            await pool.close()

        _raise_if_fetch_cancelled(job_id)
        parquet_total = sum(int(r.get("parquet_rows") or 0) for r in fetched)
        parquet_errors = [r for r in fetched if r.get("parquet_error")]
        parquet_summary = (
            f"parquet FAILED for {len(parquet_errors)} symbol(s): {parquet_errors[0]['parquet_error']}"
            if parquet_errors
            else f"parquet: {parquet_total} rows"
        )
        _jobs[job_id].update({
            "status": "done",
            "progress": 100,
            "message": (
                f"Updated {len(fetched)} existing DB {exchange.upper()} symbol(s) for {req.bar} ({parquet_summary})"
                if req.existing_only
                else f"Fetched {len(fetched)} {exchange.upper()} symbol(s) for {req.bar} ({parquet_summary})"
            ),
            "results": fetched,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    except _FetchCancelled:
        _mark_fetch_cancelled(job_id)
    except Exception as exc:
        _jobs[job_id].update({
            "status": "error",
            "message": str(exc),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
